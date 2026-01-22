from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

logger = logging.getLogger("brokerage_notes_monitor.excel")


def load_history(path: Path, sheet_name: str) -> pd.DataFrame:
    if path.exists():
        try:
            df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
            logger.info(f"Histórico existente carregado: {len(df)} linhas")
            return df
        except Exception as e:
            logger.warning(f"Erro ao ler Excel existente: {e}")
            try:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup = path.with_name(f"{path.stem}_backup_leitura_falhou_{ts}{path.suffix}")
                path.rename(backup)
                logger.warning(f"Arquivo antigo renomeado para backup: {backup}")
            except Exception as e2:
                logger.error(f"Não foi possível fazer backup do arquivo antigo: {e2}")
            return pd.DataFrame()
    return pd.DataFrame()


def backup_if_needed(path: Path) -> None:
    if not path.exists():
        return

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = path.with_name(f"{path.stem}_backup_{ts}{path.suffix}")
    path.rename(backup)
    logger.info(f"Backup criado: {backup}")


def save_history(
    df: pd.DataFrame,
    path: Path,
    sheet_name: str,
    apply_conditional_formatting: bool = True,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    df.to_excel(path, sheet_name=sheet_name, index=False, engine="openpyxl")
    logger.info(f"Histórico salvo em: {path} ({len(df)} linhas)")

    if apply_conditional_formatting:
        apply_alert_formatting(path, sheet_name)


def apply_alert_formatting(path: Path, sheet_name: str, flag_col: str = "flag_alerta_int") -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name]

    header = [cell.value for cell in ws[1]]
    if flag_col not in header:
        logger.warning(f"Coluna '{flag_col}' não encontrada para formatação condicional.")
        wb.save(path)
        return

    col_idx = header.index(flag_col) + 1
    col_letter = ws.cell(row=1, column=col_idx).column_letter

    last_row = ws.max_row
    last_col = ws.max_column
    if last_row < 2:
        wb.save(path)
        return

    start_cell = ws.cell(row=2, column=1).coordinate
    end_cell = ws.cell(row=last_row, column=last_col).coordinate
    cell_range = f"{start_cell}:{end_cell}"

    formula = f"${col_letter}2=1"

    fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    rule = FormulaRule(formula=[formula], fill=fill, stopIfTrue=False)

    ws.conditional_formatting.add(cell_range, rule)
    wb.save(path)

    logger.info("Formatação condicional aplicada (flag_alerta_int).")
